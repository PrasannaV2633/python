from flask import Flask, request, render_template, send_file, jsonify
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import requests
from bs4 import BeautifulSoup

app = Flask(__name__)

# Function to scrape product details from a given URL
def scrape_product_details(url, headers):
    try:
        # Send HTTP GET request
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, "lxml")

        # Dictionary to store extracted details
        details = {}

        # Extract product title
        title = soup.find("span", attrs={"id": "productTitle"})
        details["Title"] = title.text.strip().replace(',', '') if title else "NA"

        # Extract product price with multiple possible price spans
        price = soup.find("span", id="tp_price_block_total_price_ww")
        if price:
            # Check for multiple potential price elements for Amazon India
            price_value = price.find("span", class_="a-offscreen") or \
                          price.find("span", class_="a-price-whole")

            if price_value:
                # Clean and handle the price value
                price_text = price_value.text.strip().replace(',', '')  # Remove commas
                price_text = price_text.replace('₹', '').strip()  # Remove ₹ symbol if present
                details["Price"] = price_text
            else:
                details["Price"] = "NA"
        else:
            details["Price"] = "NA"

        return details
    except requests.RequestException as e:
        print(f"HTTP Request failed for URL {url}: {e}")
        return None
    except Exception as e:
        print(f"Error scraping {url}: {e}")
        return None


# Route to render the HTML form
@app.route('/')
def index():
    return render_template('index.html')  # Ensure the HTML file is in the 'templates' folder


# Route to handle file upload and processing
@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "No file part in the request"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    if file:
        # Save the uploaded file to the 'upload' directory
        input_file_path = os.path.join('upload', file.filename)
        file.save(input_file_path)

        try:
            # Check file extension to load accordingly
            file_extension = file.filename.rsplit('.', 1)[-1].lower()
            
            if file_extension == 'xlsx':
                df_input = pd.read_excel(input_file_path)
                df_woocommerce = pd.read_excel(input_file_path)
            elif file_extension == 'csv':
                df_input = pd.read_csv(input_file_path)
                df_woocommerce = pd.read_csv(input_file_path)
            else:
                return jsonify({"error": "Invalid file format. Only CSV and Excel files are supported."}), 400

            # Define headers for web scraping
            HEADERS = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.5938.132 Safari/537.36',
                'Accept-Language': 'en-US,en;q=0.9'
            }

            # ASIN column in the uploaded file
            asin_column = "Attribute 1 value(s)"  
            price_column = "Regular price"  
            asin_numbers = df_input[asin_column].tolist()  # Get list of ASINs from the uploaded file

            # Load the workbook for WooCommerce (for Excel files)
            if file_extension == 'xlsx':
                wb = load_workbook(input_file_path)
                ws = wb.active
            else:
                wb = None
                ws = None

            # Define fill colors for highlighting changes
            highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
            highlight_fill_updated = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green

            # Prepare an empty list to store product details
            products = []

            # Iterate over each ASIN, scrape the Amazon price, and compare
            for asin in asin_numbers:
                new_price = None
                url = f"https://www.amazon.in/dp/{asin}"

                print(f"Processing ASIN: {asin} -> {url}")
                product_details = scrape_product_details(url, HEADERS)

                if product_details:
                    amazon_price = product_details["Price"]
                    
                    # Fetch price from the WooCommerce file
                    woo_price = df_woocommerce.loc[df_woocommerce[asin_column] == asin, price_column].values

                    if len(woo_price) > 0:
                        woo_price = woo_price[0]

                        # Convert to float and compare prices
                        if amazon_price != "NA" and woo_price != "NA":
                            try:
                                amazon_price = float(amazon_price)
                                woo_price = float(woo_price)

                                # Update price if necessary (reduce by ₹10)
                                new_price = max(amazon_price - 10, 0)

                                # Update the WooCommerce price if it differs
                                if woo_price != new_price:
                                    df_woocommerce.loc[df_woocommerce[asin_column] == asin, price_column] = new_price

                                    # Highlight updated prices in the Excel file (Green)
                                    if ws:
                                        for row in ws.iter_rows():
                                            if row[40].value == asin:
                                                row[25].value = new_price
                                                row[25].fill = highlight_fill_updated
                                else:
                                    # Highlight unchanged prices in yellow
                                    if ws:
                                        for row in ws.iter_rows():
                                            if row[40].value == asin:
                                                row[25].fill = highlight_fill

                            except ValueError as e:
                                print(f"Error converting prices for ASIN {asin}: {e}")

                    # Append product details for comparison report
                    products.append([asin, product_details["Title"], amazon_price, woo_price, new_price])

            # Save the updated WooCommerce file
            updated_wc_file = "updated_wc_product_export.csv"
            if wb:
                wb.save(updated_wc_file)

            # Save a comparison report as an Excel file
            output_file_path = "amazon_vs_woocommerce_price_comparison.xlsx"
            df_output = pd.DataFrame(products, columns=["ASIN", "Product Name", "Amazon Price", "WooCommerce Price", "New Price"])
            df_output.to_excel(output_file_path, index=False)

            # Send the updated WooCommerce file back to the user
            return send_file(updated_wc_file, as_attachment=True)

        except Exception as e:
            return jsonify({"error": f"An error occurred during processing: {str(e)}"}), 500

        finally:
            # Clean up the temporary uploaded file
            if os.path.exists(input_file_path):
                os.remove(input_file_path)


# Run the Flask app
if __name__ == '__main__':
    app.run(debug=True)
